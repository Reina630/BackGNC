import mimetypes
import json
import io
import os
from datetime import datetime

from django.core.files.base import ContentFile
from django.db.models import Q
from django.http import FileResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import rest_framework as filters

# Imports pour la signature PDF
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.utils import ImageReader
from PIL import Image
# Create your views here.
from rest_framework import viewsets, status
from rest_framework import filters as rest_filters
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import (
    Courrier, PartageLog, Categorie,
    CourrierPieceJointe, ActionLog
)
from .serializer import (
    UserSimpleSerializer, 
    CourrierSerializer,
    CourrierCreateSerializer,
    CourrierUpdateSerializer,
    
    CategorieSerializer,
    ServiceSimpleSerializer
)
from users.models import User, Service
from users.permissions import IsRHOrAdmin

# ============================================================================
# VIEWSET POUR LE REGISTRE DE COURRIER
# ============================================================================

class CourrierFilter(filters.FilterSet):
    """
    Filtre personnalisé pour la recherche de courriers.
    Permet de filtrer par date, type, statut, service et recherche globale.
    """
    # Filtres de date
    date_debut = filters.DateFilter(field_name="date_reception", lookup_expr='gte')
    date_fin = filters.DateFilter(field_name="date_reception", lookup_expr='lte')
    
    # Filtre par service (accepte l'ID du service depuis la table Service)
    service = filters.NumberFilter(method='filter_by_service')
    
    # Recherche globale (sur plusieurs champs)
    search = filters.CharFilter(method='filter_search')
    
    def filter_by_service(self, queryset, name, value):
        """
        Filtrer par service en utilisant l'ID du service depuis la table Service.
        Convertit l'ID en code pour filtrer sur le champ service_concerne.
        """
        try:
            service = Service.objects.get(id=value)
            # Mapper le nom du service vers son code
            service_code = Courrier.get_service_code_from_name(service.nom)
            if service_code:
                return queryset.filter(service_concerne=service_code)
            return queryset
        except Service.DoesNotExist:
            return queryset.none()
    
    def filter_search(self, queryset, name, value):
        """
        Recherche dans plusieurs champs simultanément.
        Cherche dans : numéro de registre, objet, expéditeur, destinataire, référence
        """
        return queryset.filter(
            Q(numero_registre__icontains=value) |
            Q(objet__icontains=value) |
            Q(expediteur__icontains=value) |
            Q(destinataire__icontains=value) |
            Q(reference__icontains=value)
        )
    
    class Meta:
        model = Courrier
        fields = {
            'type_courrier': ['exact'],
            'service_concerne': ['exact'],
            'statut': ['exact'],
            'urgent': ['exact'],
        }


class CourrierViewSet(viewsets.ModelViewSet):
    """
    ViewSet complet pour gérer le registre de courrier RH.
    
    Fonctionnalités :
    - CRUD complet des courriers (RH/Admin seulement)
    - Filtrage et recherche avancée (RH/Admin seulement)
    - Export Excel du registre (RH/Admin seulement)
    - Statistiques (RH/Admin seulement)
    - Mes affectations (Tous les utilisateurs authentifiés)
    - Services disponibles (Tous les utilisateurs authentifiés)
    - Affectation par service (RH/Admin seulement)
    
    Permissions variables selon l'action
    """
    queryset = Courrier.objects.all()
    serializer_class = CourrierSerializer
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    filterset_class = CourrierFilter
    search_fields = ['numero_registre', 'objet', 'expediteur', 'destinataire', 'reference']
    ordering_fields = ['created_at', 'date_reception', 'date_envoi', 'numero_registre', 'statut']
    ordering = ['-created_at']  # Par défaut, les plus récents en premier
    
    def get_permissions(self):
        """
        Permissions variables selon l'action :
        - Actions accessibles à tous les utilisateurs authentifiés : list, retrieve, mes_affectations, services_disponibles, mes_courriers
        - Autres actions : RH et Admin seulement
        """
        if self.action in ['list', 'retrieve', 'mes_affectations', 'services_disponibles', 'mes_courriers']:
            # Actions accessibles aux utilisateurs normaux
            permission_classes = [IsAuthenticated]
        else:
            # Actions réservées aux RH et Admin
            permission_classes = [IsAuthenticated, IsRHOrAdmin]
        
        return [permission() for permission in permission_classes]
    
    def get_serializer_class(self):
        """
        Retourne le serializer approprié selon l'action.
        - Création : CourrierCreateSerializer (simplifié)
        - Mise à jour partielle : CourrierUpdateSerializer
        - Autres : CourrierSerializer (complet)
        """
        if self.action == 'create':
            return CourrierCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return CourrierUpdateSerializer
        return CourrierSerializer
    
    def get_queryset(self):
        """
        Retourne les courriers en filtrant les courriers supprimés par défaut.
        Filtre aussi selon le rôle de l'utilisateur :
        - RH/Admin : tous les courriers
        - Autres : courriers de leur service ou qui leur sont affectés
        L'action 'archives' et 'restore' peuvent accéder aux courriers supprimés.
        """
        user = self.request.user
        
        # Pour l'action archives et restore, on veut les courriers supprimés
        if self.action in ['archives', 'restore']:
            return Courrier.objects.filter(is_deleted=True)

        # Pour versions et retrieve d'un courrier archivé, inclure les courriers archivés
        if self.action in ['versions', 'retrieve', 'download', 'changer_statut', 'archives_status']:
            return Courrier.objects.filter(is_deleted=False).select_related(
                'categorie', 'enregistre_par', 'courrier_parent', 'reponse_a', 'deleted_by'
            ).prefetch_related(
                'circuits_v2', 'circuits_v2__affectations',
                'circuits_v2__affectations__destinataire',
                'circuits_v2__affectations__service',
                'affectations_v2', 'affectations_v2__destinataire',
                'affectations_v2__service',
            )
        
        # Pour les statistiques, on veut TOUS les courriers (y compris archivés) pour avoir des stats complètes
        if self.action == 'statistiques':
            queryset = Courrier.objects.filter(is_deleted=False).select_related(
                'categorie',
                'enregistre_par',
                'courrier_parent',
                'reponse_a',
                'deleted_by'
            ).prefetch_related(
                'circuits_v2',
                'circuits_v2__affectations',
                'circuits_v2__affectations__destinataire',
                'circuits_v2__affectations__service',
                'affectations_v2',
                'affectations_v2__destinataire',
                'affectations_v2__service',
            )
            return queryset
        
        # Par défaut, on ne montre que les courriers non supprimés ET non archivés
        queryset = Courrier.objects.filter(is_deleted=False).exclude(statut='archive').select_related(
            'categorie',
            'enregistre_par',
            'courrier_parent',
            'reponse_a',
            'deleted_by'
        ).prefetch_related(
            'circuits_v2',
            'circuits_v2__affectations',
            'circuits_v2__affectations__destinataire',
            'circuits_v2__affectations__service',
            'affectations_v2',
            'affectations_v2__destinataire',
            'affectations_v2__service',
        )
        
        # Filtrage selon le rôle pour list et retrieve
        if self.action in ['list', 'retrieve']:
            if user.role not in ['rh', 'admin']:
                # Utilisateurs normaux voient :
                # - Courriers de leur service (si service défini)
                # - Courriers qui leur sont affectés (nouveau système v2)
                # - Courriers qu'ils ont créés
                from django.db.models import Q
                filters = Q(enregistre_par=user) | Q(affectations_v2__destinataire=user)
                if user.service:
                    filters |= Q(service_concerne=user.service)
                queryset = queryset.filter(filters).distinct()
        
        return queryset
    
    def perform_create(self, serializer):
        """
        Enregistrer le courrier et assigner automatiquement l'utilisateur connecté.
        Calculer aussi la taille du fichier uploadé.
        Créer les pièces jointes supplémentaires si fournies.
        """
        courrier = serializer.save(enregistre_par=self.request.user)

        # Déterminer la taille du fichier principal
        if courrier.fichier:
            courrier.file_size = courrier.fichier.size
            courrier.save()

        # Traiter les pièces jointes multiples (fichiers[])
        fichiers_supplementaires = self.request.FILES.getlist('fichiers')
        for f in fichiers_supplementaires:
            ext = f.name.split('.')[-1].lower()
            if ext == 'pdf':
                ftype = 'pdf'
            elif ext in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
                ftype = 'image'
            else:
                ftype = ext
            CourrierPieceJointe.objects.create(
                courrier=courrier,
                fichier=f,
                nom_fichier=f.name,
                file_type=ftype,
                file_size=f.size,
                uploaded_by=self.request.user,
            )

        ActionLog.log_action(
            action_type='courrier_create',
            utilisateur=self.request.user,
            description=f"Courrier {courrier.numero_registre} enregistré : {courrier.objet}",
            courrier=courrier,
            request=self.request,
        )

    def perform_update(self, serializer):
        """Logger la modification d'un courrier."""
        courrier = serializer.save()
        ActionLog.log_action(
            action_type='courrier_update',
            utilisateur=self.request.user,
            description=f"Courrier {courrier.numero_registre} modifié : {courrier.objet}",
            courrier=courrier,
            request=self.request,
        )

    @action(detail=False, methods=['post'])
    def upload(self, request):
        """
        Action personnalisée pour uploader un courrier avec son fichier.
        URL : POST /api/courriers/upload/
        """
        serializer = CourrierCreateSerializer(data=request.data)
        if serializer.is_valid():
            courrier = serializer.save(enregistre_par=request.user)
            
            # Déterminer la taille du fichier
            if courrier.fichier:
                courrier.file_size = courrier.fichier.size
                # Déterminer le type de fichier
                file_extension = courrier.fichier.name.split('.')[-1].lower()
                if file_extension in ['pdf']:
                    courrier.file_type = 'pdf'
                elif file_extension in ['jpg', 'jpeg', 'png', 'gif']:
                    courrier.file_type = 'image'
                else:
                    courrier.file_type = file_extension
                courrier.save()
            
            return Response(
                CourrierSerializer(courrier).data,
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def statistiques(self, request):
        """
        Obtenir des statistiques complètes sur les courriers.
        URL : GET /api/courriers/statistiques/
        
        Retourne :
        - Total de courriers avec variation %
        - Nombre de courriers entrants/sortants
        - Répartition par statut et service
        - Courriers urgents avec détails
        - Flux de traitement (lifecycle)
        - Statistiques de versions
        - Tendances mensuelles (6 derniers mois)
        - Statistiques de partage
        - Charge de travail par service
        """
        from django.db.models import Count, Q
        from django.utils import timezone
        from datetime import timedelta
        import calendar
        
        queryset = self.get_queryset()
        now = timezone.now()
        
        # Période actuelle (30 derniers jours)
        periode_actuelle_debut = now - timedelta(days=30)
        periode_precedente_debut = now - timedelta(days=60)
        periode_precedente_fin = periode_actuelle_debut
        
        # Courriers période actuelle
        courriers_actuels = queryset.filter(created_at__gte=periode_actuelle_debut)
        total_actuel = courriers_actuels.count()
        
        # Courriers période précédente (pour comparaison)
        courriers_precedents = queryset.filter(
            created_at__gte=periode_precedente_debut,
            created_at__lt=periode_precedente_fin
        )
        total_precedent = courriers_precedents.count()
        
        # Calculer variations en %
        def calculer_variation(actuel, precedent):
            if precedent == 0:
                return 100 if actuel > 0 else 0
            return round(((actuel - precedent) / precedent) * 100, 1)
        
        # Statistiques générales avec variations
        stats = {
            'total': queryset.count(),
            'total_30j': total_actuel,
            'variation_total': calculer_variation(total_actuel, total_precedent),
            
            'entrants': queryset.filter(type_courrier='entrant').count(),
            'entrants_30j': courriers_actuels.filter(type_courrier='entrant').count(),
            'variation_entrants': calculer_variation(
                courriers_actuels.filter(type_courrier='entrant').count(),
                courriers_precedents.filter(type_courrier='entrant').count()
            ),
            
            'sortants': queryset.filter(type_courrier='sortant').count(),
            'sortants_30j': courriers_actuels.filter(type_courrier='sortant').count(),
            'variation_sortants': calculer_variation(
                courriers_actuels.filter(type_courrier='sortant').count(),
                courriers_precedents.filter(type_courrier='sortant').count()
            ),
            
            'urgents': queryset.filter(urgent=True).count(),
            'urgents_30j': courriers_actuels.filter(urgent=True).count(),
            'variation_urgents': calculer_variation(
                courriers_actuels.filter(urgent=True).count(),
                courriers_precedents.filter(urgent=True).count()
            ),
        }
        
        # Flux de traitement (Lifecycle Flow)
        # On dérive chaque étape depuis les affectations, pas depuis courrier.statut
        # car ce champ ne contient pas de valeur 'affecte'.
        from affectations.models import Affectation as AffectationV2

        # IDs de tous les courriers du queryset ayant ≥1 affectation active (non renvoyée/rejetée)
        ids_avec_affectation = set(
            AffectationV2.objects.filter(courrier__in=queryset)
            .exclude(statut__in=['renvoye', 'rejete'])
            .values_list('courrier_id', flat=True)
        )

        # "En traitement" = courriers avec ≥1 affectation en_traitement
        ids_en_traitement = set(
            AffectationV2.objects.filter(courrier__in=queryset, statut='en_traitement')
            .values_list('courrier_id', flat=True)
        )

        nb_enregistres = queryset.exclude(statut='archive').count()
        nb_affectes    = len(ids_avec_affectation)
        nb_en_traitement = len(ids_en_traitement)
        nb_traites     = queryset.filter(statut='traite').count()
        nb_archives    = queryset.filter(statut='archive').count()

        lifecycle_flow = {
            'recu': {
                'label': 'Enregistrés',
                'count': nb_enregistres,
                'color': '#dc2626'
            },
            'affecte': {
                'label': 'Affectés',
                'count': nb_affectes,
                'color': '#f59e0b'
            },
            'en_traitement': {
                'label': 'En traitement',
                'count': nb_en_traitement,
                'color': '#3b82f6'
            },
            'traite': {
                'label': 'Validés',
                'count': nb_traites,
                'color': '#10b981'
            },
            'archive': {
                'label': 'Archivés',
                'count': nb_archives,
                'color': '#6b7280'
            }
        }
        stats['lifecycle_flow'] = lifecycle_flow
        
        # Répartition par statut
        par_statut = {}
        for statut_key, statut_label in Courrier.STATUS_CHOICES:
            count = queryset.filter(statut=statut_key).count()
            par_statut[statut_key] = {
                'label': statut_label,
                'count': count
            }
        stats['par_statut'] = par_statut
        
        # Répartition par service basée sur les services réels de la base de données
        from users.models import Service as ServiceModel
        from affectations.models import Affectation as AffectationModel
        par_service = {}
        services_db = ServiceModel.objects.all()
        total_affectations_services = AffectationModel.objects.filter(service__isnull=False).count()
        for service_obj in services_db:
            count = AffectationModel.objects.filter(service=service_obj).count()
            en_traitement = AffectationModel.objects.filter(
                service=service_obj,
                statut__in=['recu', 'affecte', 'en_traitement']
            ).count()
            if count > 0:
                pourcentage = round((count / total_affectations_services * 100), 1) if total_affectations_services > 0 else 0
                par_service[service_obj.nom] = {
                    'label': service_obj.nom,
                    'count': count,
                    'en_traitement': en_traitement,
                    'pourcentage': pourcentage
                }
        stats['par_service'] = par_service
        
        # Distribution par type (pour le graphique en camembert)
        distribution_types = [
            {
                'name': 'Entrants',
                'value': stats['entrants'],
                'percentage': round((stats['entrants'] / stats['total'] * 100), 1) if stats['total'] > 0 else 0
            },
            {
                'name': 'Sortants',
                'value': stats['sortants'],
                'percentage': round((stats['sortants'] / stats['total'] * 100), 1) if stats['total'] > 0 else 0
            },
        ]
        # Ajouter internes si > 0
        count_internes = queryset.filter(type_courrier='interne').count()
        if count_internes > 0:
            distribution_types.append({
                'name': 'Internes',
                'value': count_internes,
                'percentage': round((count_internes / stats['total'] * 100), 1) if stats['total'] > 0 else 0
            })
        stats['distribution_types'] = distribution_types
        
        # Courriers urgents avec détails complets
        courriers_urgents = queryset.filter(urgent=True).exclude(
            statut__in=['traite', 'archive']
        ).order_by('-created_at')[:5]
        
        urgents_details = []
        for courrier in courriers_urgents:
            # Calculer le temps écoulé
            temps_ecoule = now - courrier.created_at
            if temps_ecoule.days > 0:
                temps_str = f"{temps_ecoule.days}j"
            else:
                heures = temps_ecoule.seconds // 3600
                temps_str = f"{heures}h"
            
            urgents_details.append({
                'id': courrier.id,
                'numero_registre': courrier.numero_registre,
                'objet': courrier.objet[:100] if courrier.objet else 'Sans objet',
                'expediteur': courrier.expediteur if hasattr(courrier, 'expediteur') else '',
                'service': dict(Courrier.SERVICE_CHOICES).get(courrier.service_concerne, 'Non défini'),
                'service_key': courrier.service_concerne,
                'statut': courrier.get_statut_display(),
                'statut_key': courrier.statut,
                'temps_ecoule': temps_str,
                'created_at': courrier.created_at.isoformat()
            })
        stats['urgents_details'] = urgents_details
        
        # Statistiques de versions
        courriers_avec_versions = queryset.filter(
            Q(courrier_parent__isnull=False) | Q(versions__isnull=False)
        ).distinct().count()
        stats['courriers_avec_versions'] = courriers_avec_versions
        stats['total_versions'] = queryset.filter(courrier_parent__isnull=False).count()
        
        # Tendances mensuelles (6 derniers mois)
        tendances = []
        for i in range(5, -1, -1):
            # Calculer le premier et dernier jour du mois
            target_month = now.month - i
            target_year = now.year
            while target_month <= 0:
                target_month += 12
                target_year -= 1
            
            # Premier jour du mois
            start_date = timezone.datetime(target_year, target_month, 1, tzinfo=now.tzinfo)
            # Dernier jour du mois
            last_day = calendar.monthrange(target_year, target_month)[1]
            end_date = timezone.datetime(target_year, target_month, last_day, 23, 59, 59, tzinfo=now.tzinfo)
            
            # Compter les courriers du mois
            count_entrants = queryset.filter(
                created_at__gte=start_date,
                created_at__lte=end_date,
                type_courrier='entrant'
            ).count()
            count_sortants = queryset.filter(
                created_at__gte=start_date,
                created_at__lte=end_date,
                type_courrier='sortant'
            ).count()
            count_total = count_entrants + count_sortants
            
            # Nom du mois en français
            mois_noms = ['', 'Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun', 'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']
            
            tendances.append({
                'mois': f"{mois_noms[target_month]} {target_year}",
                'count': count_total,
                'total': count_total,
                'entrants': count_entrants,
                'sortants': count_sortants
            })
        
        stats['tendances_mensuelles'] = tendances
        
        # Statistiques de partage (si disponibles)
        try:
            from .models import PartageLog
            stats['partages_total'] = PartageLog.objects.filter(courrier__isnull=False).count()
            stats['partages_email'] = PartageLog.objects.filter(type_partage='email').count()
            stats['partages_whatsapp'] = PartageLog.objects.filter(type_partage='whatsapp').count()
            
            # Partages cette semaine
            semaine_derniere = now - timedelta(days=7)
            stats['partages_cette_semaine'] = PartageLog.objects.filter(
                courrier__isnull=False,
                created_at__gte=semaine_derniere
            ).count()
        except:
            pass
        
        # ========================================
        # Format adapté pour le nouveau design dashboard
        # ========================================
        
        # 1. KPIs (4 cards) - format design
        recus_aujourdhui = queryset.filter(created_at__date=now.date()).count()
        recus_hier = queryset.filter(created_at__date=(now - timedelta(days=1)).date()).count()
        en_attente = queryset.filter(statut__in=['recu', 'affecte', 'en_traitement']).count()
        en_attente_avant = courriers_precedents.filter(statut__in=['recu', 'affecte', 'en_traitement']).count()

        # Urgents = affectations v2 avec niveau_urgence critique ou élevé, non terminées
        # (même source que urgentItems, donc les deux chiffres seront cohérents)
        from affectations.models import Affectation as _AffV2
        nb_urgents = _AffV2.objects.filter(
            niveau_urgence__in=['critique', 'eleve']
        ).exclude(
            statut__in=['valide', 'signe', 'rejete', 'renvoye']
        ).values('courrier_id').distinct().count()

        stats['kpis'] = [
            {
                'label': 'Total courriers',
                'value': f"{total_actuel:,}".replace(',', ' '),
                'change': f"{'+' if stats['variation_total'] >= 0 else ''}{stats['variation_total']}%",
                'positive': stats['variation_total'] >= 0,
                'color': 'bg-sky-50/80 border-sky-100'
            },
            {
                'label': 'Reçus aujourd\'hui',
                'value': str(recus_aujourdhui),
                'change': f"+{round(((recus_aujourdhui - recus_hier) / recus_hier * 100) if recus_hier > 0 else 0)}%",
                'positive': True,
                'color': 'bg-sky-50/80 border-sky-100'
            },
            {
                'label': 'En attente',
                'value': str(en_attente),
                'change': f"{'+' if calculer_variation(en_attente, en_attente_avant) >= 0 else ''}{calculer_variation(en_attente, en_attente_avant)}%",
                'positive': calculer_variation(en_attente, en_attente_avant) <= 0,
                'color': 'bg-amber-50/80 border-amber-100'
            },
            {
                'label': 'Urgents',
                'value': str(nb_urgents),
                'change': 'Élevé' if nb_urgents > 10 else ('Moyen' if nb_urgents > 0 else 'Normal'),
                'positive': False,
                'color': 'bg-red-50/80 border-red-100'
            }
        ]
        
        # 2. Lifecycle - format array (design)
        stats['lifecycle'] = [
            {
                'label': lifecycle_flow['recu']['label'],
                'count': lifecycle_flow['recu']['count'],
                'color': lifecycle_flow['recu']['color']
            },
            {
                'label': lifecycle_flow['affecte']['label'],
                'count': lifecycle_flow['affecte']['count'],
                'color': '#38bdf8'  # sky-400
            },
            {
                'label': lifecycle_flow['en_traitement']['label'],
                'count': lifecycle_flow['en_traitement']['count'],
                'color': '#1d4ed8'  # blue-700
            },
            {
                'label': lifecycle_flow['traite']['label'],
                'count': lifecycle_flow['traite']['count'],
                'color': lifecycle_flow['traite']['color']
            },
            {
                'label': lifecycle_flow['archive']['label'],
                'count': lifecycle_flow['archive']['count'],
                'color': lifecycle_flow['archive']['color']
            }
        ]
        
        # 3. Distribution par type - format design avec percentages
        stats['distribution'] = []
        total_types = stats['total'] if stats['total'] > 0 else 1
        
        # Utiliser les catégories de courrier si disponibles
        try:
            from .models import CategorieCourrier
            categories = CategorieCourrier.objects.all()[:3]
            if categories.exists():
                colors = ['#800020', '#505f76', '#c3c6d6']
                for idx, cat in enumerate(categories):
                    count = queryset.filter(categorie=cat).count()
                    percent = round((count / total_types * 100), 0)
                    stats['distribution'].append({
                        'name': cat.nom,
                        'percent': int(percent),
                        'color': colors[idx] if idx < len(colors) else '#94a3b8'
                    })
            else:
                # Fallback: utiliser les types de courrier
                stats['distribution'] = [
                    {
                        'name': 'Entrants',
                        'percent': round((stats['entrants'] / total_types * 100), 0),
                        'color': '#800020'
                    },
                    {
                        'name': 'Sortants',
                        'percent': round((stats['sortants'] / total_types * 100), 0),
                        'color': '#505f76'
                    },
                    {
                        'name': 'Internes',
                        'percent': round((count_internes / total_types * 100), 0),
                        'color': '#c3c6d6'
                    }
                ]
        except:
            stats['distribution'] = [
                {
                    'name': 'Entrants',
                    'percent': round((stats['entrants'] / total_types * 100), 0),
                    'color': '#800020'
                },
                {
                    'name': 'Sortants',
                    'percent': round((stats['sortants'] / total_types * 100), 0),
                    'color': '#505f76'
                },
                {
                    'name': 'Internes',
                    'percent': round((count_internes / total_types * 100), 0),
                    'color': '#c3c6d6'
                }
            ]
        
        # 4. urgent Items - format design (basé sur affectations critiques)
        from affectations.models import Affectation
        
        # Nouveau système : Affectation v2 avec niveau_urgence='critique' ou 'eleve'
        affectations_critiques_v2 = Affectation.objects.filter(
            niveau_urgence__in=['critique', 'eleve']
        ).exclude(
            statut__in=['valide', 'signe', 'rejete', 'renvoye']  # Exclure toutes les affectations terminées
        ).select_related('courrier', 'destinataire', 'service').order_by('-date_affectation')
        
        stats['urgentItems'] = []
        
        # Ajouter affectations du nouveau système v2
        for affectation in affectations_critiques_v2:
            # Calculer le temps écoulé depuis l'affectation
            temps_ecoule = now - affectation.date_affectation
            if temps_ecoule.days > 0:
                temps_str = f"{temps_ecoule.days}j"
            else:
                heures = temps_ecoule.seconds // 3600
                temps_str = f"{heures}h"
            
            # Créer le subtitle avec le statut et l'utilisateur
            subtitle = f"Affecté à {affectation.destinataire.get_full_name() or affectation.destinataire.username} · {temps_str}"
            
            # Déterminer le service (utiliser affectation.service si disponible, sinon service du courrier)
            if affectation.service:
                department = affectation.service.nom
            else:
                department = dict(Courrier.SERVICE_CHOICES).get(affectation.courrier.service_concerne, 'Non défini')
            
            stats['urgentItems'].append({
                'id': affectation.courrier.id,
                'affectation_id': affectation.id,
                'title': affectation.courrier.objet[:60] if affectation.courrier.objet else 'Sans objet',
                'subtitle': subtitle,
                'department': department,
                'numero_registre': affectation.courrier.numero_registre,
                'statut_affectation': affectation.get_statut_display(),
                'niveau_urgence': affectation.get_niveau_urgence_display(),
                'status': 'critique' if affectation.niveau_urgence == 'critique' else 'urgent'
            })
        
        # Limiter à 10 items et trier par date décroissante
        stats['urgentItems'] = sorted(stats['urgentItems'], key=lambda x: x.get('affectation_id', 0), reverse=True)[:10]
        
        # 5. Recent Mails - format design
        courriers_recents = queryset.order_by('-created_at')[:3]
        stats['recentMails'] = []
        
        icon_map = {
            'entrant': 'Inbox',
            'sortant': 'Send',
            'interne': 'Mail'
        }
        
        icon_color_map = {
            'entrant': 'bg-blue-50 text-blue-600',
            'sortant': 'bg-emerald-50 text-emerald-600',
            'interne': 'bg-purple-50 text-purple-600'
        }
        
        status_map = {
            'recu': 'pending',
            'affecte': 'pending',
            'en_traitement': 'in_progress',
            'traite': 'completed',
            'archive': 'completed'
        }
        
        for courrier in courriers_recents:
            # Format date relative
            diff = now - courrier.created_at
            if diff.days == 0:
                received_str = courrier.created_at.strftime('%H:%M')
            elif diff.days == 1:
                received_str = "Hier, " + courrier.created_at.strftime('%H:%M')
            else:
                received_str = courrier.created_at.strftime('%d %b, %H:%M')
            
            # Déterminer si urgent
            mail_status = status_map.get(courrier.statut, 'pending')
            if courrier.urgent:
                mail_status = 'urgent'
                
            stats['recentMails'].append({
                'id': courrier.id,
                'subject': courrier.objet[:60] if courrier.objet else 'Sans objet',
                'sender': courrier.expediteur if courrier.type_courrier == 'entrant' else courrier.destinataire,
                'received': received_str,
                'department': courrier.categorie.name if courrier.categorie else courrier.get_type_courrier_display(),
                'status': mail_status,
                'icon': icon_map.get(courrier.type_courrier, 'Mail'),
                'iconColor': icon_color_map.get(courrier.type_courrier, 'bg-slate-50 text-slate-600')
            })
        
        # 6. Service Workload - basé uniquement sur les services réels de la base de données
        from collections import defaultdict
        from users.models import Service

        charge_par_service = defaultdict(lambda: {'count': 0, 'label': '', 'en_traitement': 0})

        # Compter les affectations actives par service (Service model = source de vérité)
        affectations_actives_v2 = Affectation.objects.exclude(
            statut__in=['valide', 'signe', 'rejete', 'renvoye']
        ).select_related('service')

        for affectation in affectations_actives_v2:
            if affectation.service:
                service_nom = affectation.service.nom
                charge_par_service[service_nom]['label'] = service_nom
                charge_par_service[service_nom]['count'] += 1
                charge_par_service[service_nom]['en_traitement'] += 1
        
        # Calculer les pourcentages
        total_courriers_services = sum(s['count'] for s in charge_par_service.values())
        
        stats['serviceWorkload'] = []
        colors = ['bg-[#800020]', 'bg-emerald-500', 'bg-amber-500', 'bg-slate-500']
        
        # Trier par count décroissant et prendre top 4
        services_sorted = sorted(
            [(key, data) for key, data in charge_par_service.items()],
            key=lambda x: x[1]['count'],
            reverse=True
        )[:4]
        
        for idx, (key, service_data) in enumerate(services_sorted):
            pourcentage = round((service_data['count'] / total_courriers_services * 100), 1) if total_courriers_services > 0 else 0
            stats['serviceWorkload'].append({
                'name': service_data['label'],
                'percent': int(pourcentage),
                'color': colors[idx] if idx < len(colors) else 'bg-slate-400'
            })
        
        # 7. Weekly Trend - 7 derniers jours
        stats['weeklyTrend'] = []
        jours_fr = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
        
        for i in range(6, -1, -1):  # 7 jours en arrière
            date_jour = (now - timedelta(days=i)).date()
            count_jour = queryset.filter(created_at__date=date_jour).count()
            
            # Calculer le % relatif (max = 100%)
            jour_semaine_idx = date_jour.weekday()  # 0=Lundi, 6=Dimanche
            
            stats['weeklyTrend'].append({
                'day': jours_fr[jour_semaine_idx],
                'value': count_jour
            })
        
        # Normaliser les valeurs de weeklyTrend en pourcentages (0-100)
        max_val = max([t['value'] for t in stats['weeklyTrend']]) if stats['weeklyTrend'] else 1
        if max_val > 0:
            for trend in stats['weeklyTrend']:
                trend['value'] = round((trend['value'] / max_val) * 100)
        
        return Response(stats)
    
    @action(detail=False, methods=['get'], url_path='search-courriers')
    def search_courriers(self, request):
        """
        Rechercher des courriers pour la liste déroulante.
        URL : GET /api/courriers/search-courriers/?q=texte&type=entrant
        
        Paramètres optionnels:
        - q : texte de recherche (numero_registre, objet, expediteur, destinataire)
        - type : filtrer par type de courrier (entrant, sortant, interne)
        - exclude : ID de courrier à exclure (utile pour éviter auto-référence)
        
        Retourne une liste simplifiée de courriers avec :
        - id, numero_registre, objet, type_courrier, type_courrier_display, date_principale
        
        Limité à 50 résultats max pour les performances.
        """
        from django.db.models import Q
        
        queryset = self.get_queryset()
        
        # Recherche par texte
        q = request.query_params.get('q', '').strip()
        if q:
            queryset = queryset.filter(
                Q(numero_registre__icontains=q) |
                Q(objet__icontains=q) |
                Q(expediteur__icontains=q) |
                Q(destinataire__icontains=q) |
                Q(reference__icontains=q) |
                Q(reference_structure__icontains=q)
            )
        
        # Filtrer par type
        type_courrier = request.query_params.get('type', '').strip()
        if type_courrier in ['entrant', 'sortant', 'interne']:
            queryset = queryset.filter(type_courrier=type_courrier)
        
        # Exclure un courrier spécifique (pour éviter auto-référence)
        exclude_id = request.query_params.get('exclude', '').strip()
        if exclude_id and exclude_id.isdigit():
            queryset = queryset.exclude(id=int(exclude_id))
        
        # Limiter à 50 résultats et trier par date décroissante
        courriers = queryset.order_by('-created_at')[:50]
        
        # Construire la réponse simplifiée
        results = []
        for courrier in courriers:
            results.append({
                'id': courrier.id,
                'numero_registre': courrier.numero_registre,
                'objet': courrier.objet,
                'type_courrier': courrier.type_courrier,
                'type_courrier_display': courrier.get_type_courrier_display(),
                'date_principale': courrier.get_date_principale().isoformat() if courrier.get_date_principale() else None,
                'expediteur': courrier.expediteur,
                'destinataire': courrier.destinataire,
            })
        
        return Response(results)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Exporter le registre de courrier au format Excel.
        URL : GET /api/courriers/export_excel/
        Paramètre optionnel : fields=numero_registre,type_courrier,...
        
        Génère un fichier Excel avec les colonnes sélectionnées.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse
        from datetime import datetime
        
        # Appliquer les filtres de la requête
        queryset = self.filter_queryset(self.get_queryset())

        # Filtres supplémentaires non couverts par filter_queryset
        concerne = request.query_params.get('concerne', '').strip()
        expediteur = request.query_params.get('expediteur', '').strip()
        destinataire = request.query_params.get('destinataire', '').strip()
        if concerne:
            queryset = queryset.filter(
                Q(expediteur__icontains=concerne) | Q(destinataire__icontains=concerne)
            )
        if expediteur:
            queryset = queryset.filter(expediteur__icontains=expediteur)
        if destinataire:
            queryset = queryset.filter(destinataire__icontains=destinataire)

        # Définition de toutes les colonnes disponibles : clé → (label, getter)
        ALL_COLUMNS = [
            ('numero_registre',  "N° Registre",           lambda c: c.numero_registre),
            ('type_courrier',    "Type",                   lambda c: c.get_type_courrier_display()),
            ('date_reception',   "Date Réception",         lambda c: c.date_reception.strftime('%d/%m/%Y') if c.date_reception else ''),
            ('mode_reception',   "Mode de réception",      lambda c: c.get_mode_reception_display() if c.mode_reception else ''),
            ('date_envoi',       "Date Envoi",             lambda c: c.date_envoi.strftime('%d/%m/%Y') if c.date_envoi else ''),
            ('mode_envoi',       "Mode d'envoi",           lambda c: c.get_mode_envoi_display() if c.mode_envoi else ''),
            ('expediteur',       "Expéditeur",             lambda c: c.expediteur),
            ('destinataire',     "Destinataire",           lambda c: c.destinataire),
            ('objet',            "Objet",                  lambda c: c.objet),
            ('reference',        "Référence",              lambda c: c.reference),
            ('categorie',        "Catégorie",              lambda c: c.categorie.nom if c.categorie else ''),
            ('service_concerne', "Service Concerné",       lambda c: c.get_service_concerne_display() if c.service_concerne else ''),
            ('statut',           "Statut",                 lambda c: c.get_statut_display()),
            ('urgent',           "Urgent",                 lambda c: 'Oui' if c.urgent else 'Non'),
            ('notes',            "Notes",                  lambda c: c.notes),
            ('enregistre_par',   "Enregistré par",         lambda c: c.enregistre_par.username if c.enregistre_par else ''),
            ('created_at',       "Date d'enregistrement",  lambda c: c.created_at.strftime('%d/%m/%Y %H:%M')),
        ]

        # Filtrer les colonnes selon le param ?fields= (si fourni)
        requested = request.query_params.get('fields', '')
        if requested:
            requested_keys = [f.strip() for f in requested.split(',') if f.strip()]
            key_order = {k: i for i, k in enumerate(requested_keys)}
            columns = [col for col in ALL_COLUMNS if col[0] in requested_keys]
            columns.sort(key=lambda col: key_order.get(col[0], 999))
        else:
            # Par défaut : toutes les colonnes sauf mode_reception, mode_envoi, categorie, urgent
            default_keys = {'numero_registre','type_courrier','date_reception','date_envoi',
                            'expediteur','destinataire','objet','reference',
                            'service_concerne','statut','notes','enregistre_par','created_at'}
            columns = [col for col in ALL_COLUMNS if col[0] in default_keys]

        # Créer le workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registre de Courrier"
        
        # Styles pour l'en-tête
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )
        
        # Écrire les en-têtes
        for col_num, (_, label, _getter) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Écrire les données
        for row_num, courrier in enumerate(queryset, 2):
            for col_num, (_key, _label, getter) in enumerate(columns, 1):
                try:
                    value = getter(courrier)
                except Exception:
                    value = ''
                ws.cell(row=row_num, column=col_num, value=value).border = thin_border
        
        # Ajuster automatiquement la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            # Limiter la largeur maximale à 50 caractères
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Figer la première ligne (en-têtes)
        ws.freeze_panes = 'A2'
        
        # Préparer la réponse HTTP avec le fichier Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nom du fichier avec date et heure
        filename = f"registre_courrier_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le workbook dans la réponse
        wb.save(response)
        return response
    
    @action(detail=True, methods=['patch'])
    def changer_statut(self, request, pk=None):
        """
        Changer le statut d'un courrier.
        URL : PATCH /api/courriers/{id}/changer_statut/
        Body : {"statut": "traite"} (ou "recu", "en_traitement", "archive")
        """
        courrier = self.get_object()
        ancien_statut = courrier.statut
        nouveau_statut = request.data.get('statut')
        
        # Vérifier que le statut est valide
        statuts_valides = [choice[0] for choice in Courrier.STATUS_CHOICES]
        if nouveau_statut not in statuts_valides:
            return Response(
                {
                    "error": "Statut invalide",
                    "statuts_valides": dict(Courrier.STATUS_CHOICES)
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Mettre à jour le statut
        courrier.statut = nouveau_statut
        courrier.save()
        
        # Si le statut passe à "traité", notifier les personnes concernées
        if nouveau_statut == 'traite' and ancien_statut != 'traite':
            from users.utils import creer_notification
            
            # Récupérer tous ceux qui ont affecté ce courrier (nouveau système v2)
            affecteurs = courrier.affectations_v2.filter(affecte_par__isnull=False).values_list('affecte_par', flat=True).distinct()
            
            for affecteur_id in affecteurs:
                try:
                    creer_notification(
                        utilisateur=affecteur_id,
                        type_notif='courrier_affecte',
                        titre=f'Courrier traité: {courrier.numero_registre}',
                        message=f'Le courrier "{courrier.objet}" a été marqué comme traité par {request.user.get_full_name() or request.user.username}.',
                        courrier_id=courrier.id,
                    )
                except Exception as e:
                    print(f"Erreur lors de la création de notification: {e}")
        
        return Response({
            "message": f"Statut mis à jour : {courrier.get_statut_display()}",
            "courrier": CourrierSerializer(courrier).data
        })
    
    @action(detail=True, methods=['post'])
    def toggle_urgent(self, request, pk=None):
        """
        Marquer/Démarquer un courrier comme urgent.
        URL : POST /api/courriers/{id}/toggle_urgent/
        """
        courrier = self.get_object()
        
        # Basculer l'état urgent
        courrier.urgent = not courrier.urgent
        courrier.save()
        
        return Response({
            "message": f"Courrier {'marqué comme urgent' if courrier.urgent else 'retiré des urgents'}",
            "urgent": courrier.urgent,
            "courrier": CourrierSerializer(courrier).data
        })
    
    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """
        Télécharger le fichier scanné d'un courrier.
        URL : GET /api/courriers/{id}/download/
        """
        courrier = self.get_object()
        
        if not courrier.fichier:
            return Response(
                {"error": "Aucun fichier associé à ce courrier"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Préparer la réponse avec le fichier
        response = FileResponse(courrier.fichier.open('rb'), content_type='application/octet-stream')
        filename = f"{courrier.numero_registre}_{courrier.fichier.name.split('/')[-1]}"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @action(detail=True, methods=['post'])
    def creer_version(self, request, pk=None):
        """
        Créer une nouvelle version d'un courrier existant.
        URL : POST /api/courriers/{id}/creer_version/
        Body (multipart/form-data) : {
            "fichier": <file>,
            "notes": "Notes sur cette version (optionnel)"
        }
        """
        courrier = self.get_object()
        
        # Vérifier qu'un fichier est fourni
        if 'fichier' not in request.FILES:
            return Response(
                {"error": "Le fichier est obligatoire pour créer une nouvelle version"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        fichier = request.FILES['fichier']
        notes = request.data.get('notes', '')
        
        try:
            # Créer la nouvelle version
            nouvelle_version = courrier.creer_nouvelle_version(
                fichier=fichier,
                notes=notes,
                enregistre_par=request.user
            )
            
            # Calculer la taille du fichier
            if nouvelle_version.fichier:
                nouvelle_version.file_size = nouvelle_version.fichier.size
                # Déterminer le type de fichier
                file_extension = nouvelle_version.fichier.name.split('.')[-1].lower()
                if file_extension in ['pdf']:
                    nouvelle_version.file_type = 'pdf'
                elif file_extension in ['jpg', 'jpeg', 'png', 'gif']:
                    nouvelle_version.file_type = 'image'
                else:
                    nouvelle_version.file_type = file_extension
                nouvelle_version.save()
            
            return Response({
                "message": f"Nouvelle version créée : {nouvelle_version.get_version_label()}",
                "version": CourrierSerializer(nouvelle_version).data
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {"error": f"Erreur lors de la création de la version : {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def versions(self, request, pk=None):
        """
        Récupérer toutes les versions d'un courrier.
        URL : GET /api/courriers/{id}/versions/
        
        Retourne toutes les versions du courrier (incluant lui-même et ses versions)
        """
        courrier = self.get_object()
        
        # Récupérer toutes les versions
        toutes_versions = courrier.get_toutes_versions()
        
        # Ajouter le courrier parent si ce n'est pas déjà une version
        if not courrier.courrier_parent:
            # Créer une liste avec le parent et ses versions
            versions_list = [courrier] + list(toutes_versions)
        else:
            # Si c'est une version, récupérer le parent et toutes les versions
            parent = courrier.courrier_parent
            versions_list = [parent] + list(parent.versions.all().order_by('version_number'))
        
        # Sérialiser toutes les versions
        serializer = CourrierSerializer(versions_list, many=True, context={'request': request})
        
        return Response({
            "nombre_versions": len(versions_list),
            "version_actuelle": courrier.get_version_actuelle().version_number if courrier.get_version_actuelle() else None,
            "versions": serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def affecter_service(self, request, pk=None):
        """
        Affecter un courrier à tous les utilisateurs d'un service.
        Crée un Circuit simultané + une Affectation par utilisateur du service
        (table affectations.Affectation — seule table d'affectation de référence).

        URL : POST /api/courriers/{id}/affecter_service/
        Body JSON :
        {
            "service_id": 1,
            "note": "...",
            "niveau_urgence": "normal|faible|eleve|critique",
            "date_echeance": "2026-03-20",
            "action_requise": "informatif|a_signer|..."
        }
        """
        from affectations.models import Circuit, Affectation
        from users.models import Notification

        try:
            courrier = self.get_object()
            service_id = request.data.get('service_id')
            note = request.data.get('note', '')
            niveau_urgence = request.data.get('niveau_urgence', 'normal')
            date_echeance = request.data.get('date_echeance') or None
            action_requise = request.data.get('action_requise', 'informatif')

            if not service_id:
                return Response(
                    {'error': 'service_id est requis'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                service = Service.objects.get(id=service_id)
            except Service.DoesNotExist:
                return Response(
                    {'error': 'Service introuvable'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Récupérer les utilisateurs actifs du service
            utilisateurs_service = User.objects.filter(service=service, is_active=True)

            if not utilisateurs_service.exists():
                return Response(
                    {'error': f'Aucun utilisateur actif trouvé dans le service {service.nom}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Créer un circuit simultané pour ce courrier
            circuit = Circuit.objects.create(
                courrier=courrier,
                type_circuit='simultane',
                titre=f'Affectation au service {service.nom}',
                cree_par=request.user,
            )

            # Créer une affectation par utilisateur du service
            affectations_creees = []
            for utilisateur in utilisateurs_service:
                affectation = Affectation.objects.create(
                    circuit=circuit,
                    courrier=courrier,
                    destinataire=utilisateur,
                    service=service,
                    affecte_par=request.user,
                    action_requise=action_requise,
                    niveau_urgence=niveau_urgence,
                    date_echeance=date_echeance,
                    note_instruction=note,
                    etape_numero=1,
                    statut='distribue',
                )
                affectations_creees.append(affectation)

                # Notification
                Notification.objects.create(
                    utilisateur=utilisateur,
                    type='courrier_affecte',
                    titre=f'Nouveau courrier affecté : {courrier.numero_registre}',
                    message=(
                        f'Le courrier "{courrier.objet}" a été affecté à votre service '
                        f'({service.nom}). Action requise : {affectation.get_action_requise_display()}'
                    ),
                    courrier_id=courrier.id,
                )

            # Mettre à jour le statut et le service du courrier
            if courrier.statut == 'recu':
                courrier.statut = 'en_traitement'
            service_code = Courrier.get_service_code_from_name(service.nom)
            courrier.service_concerne = service_code
            courrier.save()

            ActionLog.log_action(
                action_type='affectation_create',
                utilisateur=request.user,
                description=(
                    f"Courrier {courrier.numero_registre} affecté au service {service.nom} "
                    f"via circuit #{circuit.id} ({len(affectations_creees)} affectation(s))"
                ),
                courrier=courrier,
                request=request,
            )

            return Response({
                'message': f'Courrier affecté à {len(affectations_creees)} utilisateur(s) du service {service.nom}',
                'circuit_id': circuit.id,
                'service_nom': service.nom,
                'service_code': service_code,
                'utilisateurs_affectes': len(affectations_creees),
                'courrier_numero': courrier.numero_registre,
                'courrier_statut': courrier.statut,
                'courrier_service_concerne': courrier.service_concerne,
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    

    @action(detail=False, methods=['get'])
    def mes_courriers(self, request):
        """
        Récupérer les courriers affectés à l'utilisateur connecté.
        URL : GET /api/courriers/mes_courriers/

        Utilise uniquement affectations.Affectation (table de référence).
        """
        from django.db.models import Q
        from affectations.models import Affectation

        # IDs des courriers affectés au user (nouveau système, seule source de vérité)
        courriers_ids = set(
            Affectation.objects
            .filter(destinataire=request.user)
            .exclude(statut='renvoye')
            .values_list('courrier_id', flat=True)
        )

        # Récupérer les courriers correspondants
        courriers = Courrier.objects.filter(id__in=courriers_ids).order_by('-created_at')

        # Filtres optionnels
        search = request.query_params.get('search')
        if search:
            courriers = courriers.filter(
                Q(numero_registre__icontains=search) |
                Q(objet__icontains=search) |
                Q(expediteur__icontains=search) |
                Q(destinataire__icontains=search)
            )

        ordering = request.query_params.get('ordering')
        if ordering:
            courriers = courriers.order_by(ordering)

        serializer = CourrierSerializer(courriers, many=True)
        return Response(serializer.data)
    
    def destroy(self, request, *args, **kwargs):
        """
        Archiver (soft delete) un courrier au lieu de le supprimer complètement.
        URL : DELETE /api/courriers/{id}/
        """
        courrier = self.get_object()
        
        # Archiver le courrier (soft delete)
        courrier.soft_delete(request.user)
        ActionLog.log_action(
            action_type='courrier_archive',
            utilisateur=request.user,
            description=f"Courrier {courrier.numero_registre} archivé : {courrier.objet}",
            courrier=courrier,
            request=request,
        )
        return Response({
            "message": "Courrier archivé avec succès"
        }, status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=False, methods=['get'])
    def archives(self, request):
        """
        Récupérer tous les courriers archivés (supprimés) accessibles par l'utilisateur.
        URL : GET /api/courriers/archives/
        """
        # Récupérer tous les courriers supprimés
        user = request.user
        
        if user.role == 'admin' or user.role == 'rh':
            # Les admins et RH voient tous les courriers archivés
            archived_courriers = Courrier.objects.filter(is_deleted=True)
        else:
            # Les utilisateurs ne voient que leurs propres courriers archivés
            archived_courriers = Courrier.objects.filter(is_deleted=True, enregistre_par=user)
        
        # Appliquer les filtres de recherche si nécessaire
        search_query = request.query_params.get('search', None)
        if search_query:
            archived_courriers = archived_courriers.filter(
                Q(numero_registre__icontains=search_query) |
                Q(objet__icontains=search_query) |
                Q(expediteur__icontains=search_query) |
                Q(destinataire__icontains=search_query)
            )
        
        # Trier par date de suppression (plus récent en premier)
        archived_courriers = archived_courriers.order_by('-deleted_at')
        
        serializer = self.get_serializer(archived_courriers, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        """
        Restaurer un courrier archivé.
        URL : POST /api/courriers/{id}/restore/
        """
        # Récupérer le courrier même s'il est supprimé
        try:
            courrier = Courrier.objects.get(pk=pk)
        except Courrier.DoesNotExist:
            return Response(
                {"error": "Courrier non trouvé"},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Vérifier que le courrier est bien archivé
        if not courrier.is_deleted:
            return Response(
                {"error": "Ce courrier n'est pas archivé"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Restaurer le courrier
        courrier.restore()
        ActionLog.log_action(
            action_type='courrier_restore',
            utilisateur=request.user,
            description=f"Courrier {courrier.numero_registre} restauré depuis les archives.",
            courrier=courrier,
            request=request,
        )
        serializer = self.get_serializer(courrier)
        return Response({
            "message": "Courrier restauré avec succès",
            "courrier": serializer.data
        })
    
    @action(detail=False, methods=['get'], url_path='archives-status')
    def archives_status(self, request):
        """
        Récupérer tous les courriers avec statut='archive' (courriers traités et classés).
        URL : GET /api/courriers/archives-status/
        """
        user = request.user
        
        # Récupérer les courriers archivés (statut='archive') et non supprimés
        if user.role == 'admin' or user.role == 'rh':
            # Les admins et RH voient tous les courriers archivés
            archived_courriers = Courrier.objects.filter(statut='archive', is_deleted=False)
        else:
            # Les utilisateurs ne voient que leurs propres courriers archivés
            archived_courriers = Courrier.objects.filter(
                statut='archive', 
                is_deleted=False, 
                enregistre_par=user
            )
        
        # Appliquer les filtres de recherche si nécessaire
        search_query = request.query_params.get('search', None)
        if search_query:
            archived_courriers = archived_courriers.filter(
                Q(numero_registre__icontains=search_query) |
                Q(objet__icontains=search_query) |
                Q(expediteur__icontains=search_query) |
                Q(destinataire__icontains=search_query)
            )
        
        # Trier par date de création (plus récent en premier)
        ordering = request.query_params.get('ordering', '-created_at')
        archived_courriers = archived_courriers.order_by(ordering)
        
        serializer = self.get_serializer(archived_courriers, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def versions(self, request, pk=None):
        """
        Récupérer toutes les versions d'un courrier.
        URL : GET /api/courriers/{id}/versions/
        """
        courrier = self.get_object()
        
        # Récupérer toutes les versions de ce courrier
        if courrier.courrier_parent:
            # Si c'est une version, récupérer toutes les versions du parent
            parent = courrier.courrier_parent
            versions = Courrier.objects.filter(
                Q(id=parent.id) | Q(courrier_parent=parent)
            ).order_by('version_number')
        else:
            # Si c'est le parent, récupérer lui-même et toutes ses versions
            versions = Courrier.objects.filter(
                Q(id=courrier.id) | Q(courrier_parent=courrier)
            ).order_by('version_number')
        
        serializer = self.get_serializer(versions, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def services_disponibles(self, request):
        """
        Liste des services disponibles pour l'affectation.
        URL : GET /api/courriers/services_disponibles/
        """
        services = Service.objects.all().order_by('nom')
        serializer = ServiceSimpleSerializer(services, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='pieces_jointes')
    def ajouter_piece_jointe(self, request, pk=None):
        """
        Ajouter une ou plusieurs pièces jointes à un courrier existant.
        URL : POST /api/courriers/{id}/pieces_jointes/
        """
        courrier = self.get_object()
        fichiers = request.FILES.getlist('fichiers')
        if not fichiers:
            return Response({'error': 'Aucun fichier fourni.'}, status=status.HTTP_400_BAD_REQUEST)

        created = []
        for f in fichiers:
            ext = f.name.split('.')[-1].lower()
            if ext == 'pdf':
                ftype = 'pdf'
            elif ext in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
                ftype = 'image'
            else:
                ftype = ext
            pj = CourrierPieceJointe.objects.create(
                courrier=courrier,
                fichier=f,
                nom_fichier=f.name,
                file_type=ftype,
                file_size=f.size,
                uploaded_by=request.user,
            )
            created.append(pj)

        from .serializer import CourrierPieceJointeSerializer
        serializer = CourrierPieceJointeSerializer(created, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['delete'], url_path=r'pieces_jointes/(?P<pj_id>\d+)')
    def supprimer_piece_jointe(self, request, pk=None, pj_id=None):
        """
        Supprimer une pièce jointe d'un courrier.
        URL : DELETE /api/courriers/{id}/pieces_jointes/{pj_id}/
        """
        courrier = self.get_object()
        try:
            pj = CourrierPieceJointe.objects.get(id=pj_id, courrier=courrier)
        except CourrierPieceJointe.DoesNotExist:
            return Response({'error': 'Pièce jointe introuvable.'}, status=status.HTTP_404_NOT_FOUND)

        pj.fichier.delete(save=False)  # Supprimer le fichier physique
        pj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ============================================================================
# VIEWSET POUR LES CATÉGORIES DE COURRIER
# ============================================================================

class CategorieViewSet(viewsets.ModelViewSet):
    """
    ViewSet pour gérer les catégories de courriers.
    Permet de lister, créer, modifier et supprimer des catégories.
    """
    queryset = Categorie.objects.all()
    serializer_class = CategorieSerializer
    permission_classes = [IsAuthenticated]
    
    # Filtrage et recherche
    filter_backends = [DjangoFilterBackend, rest_filters.SearchFilter, rest_filters.OrderingFilter]
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']
    
    @action(detail=False, methods=['post'])
    def get_or_create(self, request):
        """
        Récupère une catégorie existante par son nom ou en crée une nouvelle.
        URL : POST /api/categories/get_or_create/
        Body : { "name": "Devis" }
        """
        name = request.data.get('name', '').strip()
        
        if not name:
            return Response(
                {"error": "Le nom de la catégorie est obligatoire"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Chercher ou créer la catégorie
        categorie, created = Categorie.objects.get_or_create(
            name=name,
            defaults={'description': request.data.get('description', '')}
        )
        
        serializer = self.get_serializer(categorie)
        status_code = status.HTTP_201_CREATED if created else status.HTTP_200_OK
        
        return Response(serializer.data, status=status_code)


# ============================================================================
# UTILITAIRES POUR LA SIGNATURE ÉLECTRONIQUE
# ============================================================================

def appliquer_signature_pdf(pdf_path, signature_path, position_x, position_y, largeur, hauteur, page_height=1200):
    """
    Applique une signature électronique sur un PDF.
    
    Args:
        pdf_path: Chemin vers le PDF original
        signature_path: Chemin vers l'image de signature
        position_x: Position X (en pixels frontend)
        position_y: Position Y (en pixels frontend)
        largeur: Largeur de la signature (en pixels)
        hauteur: Hauteur de la signature (en pixels)
        page_height: Hauteur de la zone d'affichage frontend (défaut: 1200px)
    
    Returns:
        BytesIO contenant le PDF signé
    """
    import os
    
    # Vérifier que le fichier de signature existe
    if not os.path.exists(signature_path):
        raise FileNotFoundError(f"Le fichier de signature n'existe pas : {signature_path}")
    
    # Lire le PDF original pour obtenir sa vraie taille
    existing_pdf = PdfReader(open(pdf_path, "rb"))
    first_page = existing_pdf.pages[0]
    
    # Obtenir la taille réelle de la page en points
    page_box = first_page.mediabox
    pdf_width = float(page_box.width)
    pdf_height = float(page_box.height)
    
    print(f"[appliquer_signature] PDF size: {pdf_width} x {pdf_height} points")
    print(f"[appliquer_signature] Frontend position: ({position_x}, {position_y}), size: ({largeur}, {hauteur})")
    print(f"[appliquer_signature] Frontend container height: {page_height}px")
    
    # Facteur de conversion : ratio entre hauteur PDF réelle et hauteur du conteneur frontend
    scale_factor = pdf_height / page_height
    
    print(f"[appliquer_signature] Scale factor: {scale_factor}")
    
    # Convertir position et dimensions
    # Frontend: origine en haut-gauche (0,0)
    # PDF: origine en bas-gauche (0,0)
    x_pdf = position_x * scale_factor
    y_pdf = pdf_height - (position_y * scale_factor) - (hauteur * scale_factor)
    w_pdf = largeur * scale_factor
    h_pdf = hauteur * scale_factor
    
    print(f"[appliquer_signature] PDF coordinates: ({x_pdf}, {y_pdf}), size: ({w_pdf}, {h_pdf})")
    
    # Créer un PDF temporaire avec juste la signature
    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=(pdf_width, pdf_height))
    
    # Ajouter l'image de signature
    try:
        signature_abs_path = os.path.abspath(signature_path)
        img = ImageReader(signature_abs_path)
        can.drawImage(img, x_pdf, y_pdf, width=w_pdf, height=h_pdf, mask='auto')
    except Exception as e:
        print(f"Erreur lors de l'ajout de l'image: {e}")
        raise Exception(f"Impossible de charger l'image de signature : {str(e)}")
    
    can.save()
    packet.seek(0)
    
    # Lire le PDF de signature
    signature_pdf = PdfReader(packet)
    
    # Créer le PDF de sortie
    output = PdfWriter()
    
    # Fusionner la signature avec la première page
    page = existing_pdf.pages[0]
    page.merge_page(signature_pdf.pages[0])
    output.add_page(page)
    
    # Ajouter les autres pages sans modification
    for i in range(1, len(existing_pdf.pages)):
        output.add_page(existing_pdf.pages[i])
    
    # Écrire dans un BytesIO
    output_stream = io.BytesIO()
    output.write(output_stream)
    output_stream.seek(0)
    
    return output_stream
   